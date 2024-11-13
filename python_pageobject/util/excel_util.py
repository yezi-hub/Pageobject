import openpyxl
import time
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle, Border, Side, PatternFill
import traceback

class Excel:

    def __init__(self,excel_file_path):
        if not os.path.exists(excel_file_path):
            self.excel_file_path = None
            self.wb = None
            self.sheet = None
        else:
            self.excel_file_path = excel_file_path
            self.wb = load_workbook(self.excel_file_path)
            self.sheet = None

    #获取excel对象的文件路径
    def get_excel_file_path(self):
        return self.excel_file_path

    #设置excel对象的文件路径
    def set_excel_file_path(self,excel_file_path):
        if not os.path.exists(excel_file_path):
            self.excel_file_path = None
            self.wb = None
            self.sheet = None
        else:
            self.excel_file_path = excel_file_path
            self.wb = load_workbook(self.excel_file_path)
            self.sheet = None

    #获取excel文件的所有sheet名称，返回的是列表
    def get_sheet_names(self):
        if self.wb:
            return self.wb.sheetnames
        else:
            return []

    #使用字符串创建一个sheet
    def create_sheet_by_name(self,sheet_name):
        if sheet_name in self.get_sheet_names():
            print("sheet名称 %s 已经存在了，无需创建" %sheet_name)
            return None
        else:
            self.wb.create_sheet(sheet_name)

    #获取当前设定sheet的sheet名称
    def get_current_sheet_name(self):
        if self.sheet:
            return self.sheet.title
        else:
            return None

    #设置当前使用sheet
    def set_sheet(self,sheet_name):
        if sheet_name in self.get_sheet_names():
            self.sheet = self.wb[sheet_name]
        else:
            print("要设定的sheet名称‘%s’不存在" %sheet_name)
            self.sheet  = None

    #获得所有的行对象
    def get_all_row_objects(self):
        if not self.sheet:
            return []
        rows = []
        for row in self.sheet.rows:
            rows.append(row)
        return rows

    #获取指定行的对象
    def get_one_row(self,row_no):#行号从1开始
        if not isinstance(row_no,int):
            print("%s 不是整数！无法获取行" %row_no)
            return None
        if row_no>=1 and row_no<=len(self.get_all_row_objects()):
            return self.get_all_row_objects()[row_no-1]
        else:
            print("您输入的行号 %s 不存在！" %row_no)
            return None

    #获取所有的单元格对象
    def get_all_cell_objects(self):
        if not self.get_all_row_objects():
            return []
        all_cells = []
        for row in self.get_all_row_objects():
            row_cells = []
            for cell in row:
                row_cells.append(cell)
            all_cells.append(row_cells)
        return all_cells

    #获取所有单元格的值，返回的是个二维列表
    def get_all_cell_values(self):
        all_cell_value = []
        if not self.get_all_cell_objects():
            return []

        for row in self.get_all_cell_objects():
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            all_cell_value.append(row_values)

        return all_cell_value

    #获取某一行的值
    def get_one_row_value(self,row_no):
        if not isinstance(row_no,int):
            print("%s 不是整数！无法获取行的内容" %row_no)
            return []
        values = []
        if row_no>=1 and row_no<=len(self.get_all_row_objects()):
            for cell in  self.get_all_row_objects()[row_no-1]:
                values.append(cell.value)
            return values
        else:
            print("您输入的行号 %s 不存在！" %row_no)
            return []

    # 获取所有列的对象
    def get_all_col_objects(self):
        if not self.sheet:
            print("没有设定sheet，无法读取列对象")
            return []
        all_cols = []
        for col in self.sheet.columns:
            all_cols.append(col)
        return all_cols

    #获取所有列值
    def get_all_col_values(self):
        if not self.get_all_col_objects():
            return []
        all_col_values = []
        for col in self.get_all_col_objects():
            col_values = []
            for cell in col:
                col_values.append(cell.value)
            all_col_values.append(col_values)
        return all_col_values

    #获取某一列对象，列号从1开始
    def get_one_col_object(self, col_no):
        if not isinstance(col_no, int):
            print("列号 %s 不是一个整数，无法获取列对象" % col_no)
            return []
        if col_no >= 1 and col_no <= len(self.get_all_col_objects()):
            return self.get_all_col_objects()[col_no - 1]
        else:
            print("列号 %s 不在有效范围内！" % col_no)

    #获取某一列值,，列号从1开始
    def get_one_col_values(self,col_no):
        if not isinstance(col_no, int):
            print("列号 %s 不是一个整数，无法获取列值" % col_no)
            return []
        col_values = []
        if col_no >= 1 and col_no <= len(self.get_all_col_objects()):
            for cell in self.get_all_col_objects()[col_no - 1]:
                col_values.append(cell.value)
            return col_values
        else:
            print("列号 %s 不在有效范围内！" % col_no)

    #获取某个单元格的值
    # 读取单元格的值
    def get_cell_value(self, row_no, col_no):
        if not isinstance(row_no, int):
            print("行号 %s 不是整数无法获取单元格的值！" % row_no)
            return None
        if not isinstance(col_no, int):
            print("列号 %s 不是整数无法获取单元格的值！" % col_no)
            return None

        values = self.get_all_cell_values()  # 获取所有格的值
        if row_no >= 1 and row_no <= len(self.get_all_cell_values()):
            if col_no >= 1 and col_no <= len(self.get_all_cell_values()[0]):
                return self.get_all_cell_values()[row_no - 1][col_no - 1]
            else:
                print("列号 %s 不在有效范围无法获取单元格的值！" % col_no)
                return None
        else:
            print("行号 %s 不在有效范围无法获取单元格的值！" % row_no)
            return None

    # 追加在excel的最后写入多行，lines参数是个2维列表
    def write_lines(self,lines,head_color=None):
        if not self.sheet:
            print("当前sheet没有设定，无法写入多行内容！")
            return False
        for i in  range(len(lines)):
            if i==0 and head_color:
                self.write_one_line(lines[i],pattern=head_color)
                continue
            self.write_one_line(lines[i])
        self.save()

    #追加写入一行数据
    def write_one_line(self,line,font=None,color = None,pattern=None):
        if not self.sheet:
            print("当前sheet没有设定，无法写入行内容！")
            return False
        row_no = self.sheet.max_row+1
        for col_no in range(len(line)):
            #self.sheet.cell(row = row_no,column=col_no+1).value = line[col_no]
            self.write_cell(row_no = row_no,col_no=col_no+1,value=line[col_no],font=font,color=color,pattern=pattern)

    # 写入某个单元格内容
    def write_cell(self, row_no, col_no, value, font=None, color=None, pattern=None):
        if not isinstance(row_no, int):
            print("行号 %s 不是整数无法写入单元格的值！" % row_no)
            return False
        if not isinstance(col_no, int):
            print("列号 %s 不是整数无法写入单元格的值！" % col_no)
            return False

        if not self.sheet:
            print("当前sheet没有设定，无法写入单元格内容！")
            return False

        bd = Side(style='thin', color="000000")
        border = Border(left=bd, top=bd, right=bd, bottom=bd)
        try:
            self.sheet.cell(row=row_no, column=col_no, value=value)
            self.sheet.cell(row=row_no, column=col_no).border = border
            ft = Font()#字体对象
            self.sheet.cell(row=row_no, column=col_no).font = ft
            if font:#字体
                ft.name = font
            if color:
                ft.color = color
            if pattern:
                fill = PatternFill(fill_type="solid",
                                       start_color=pattern,
                                       end_color=pattern)
                self.sheet.cell(row=row_no, column=col_no).fill = fill
            if "失败" in str(value) or "fail" in str(value).lower():
                fill = PatternFill(fill_type="solid",
                                   start_color="FF0000",
                                   end_color="FF0000")
                self.sheet.cell(row=row_no, column=col_no).fill = fill

            return True
        except Exception as e:
            traceback.print_exc()
            print("写入单元格内容时候发生异常，信息：%s" % e)
            return False


    #保存对excel文件的修改
    def save(self):
        if self.wb:
            self.wb.save(self.excel_file_path)

if __name__ =="__main__":
    #excel = Excel("e:\\aaaa.xlsx")
    #print(excel.get_excel_file_path())
    #excel = Excel("e:\\a.xlsx")
    #print(excel.get_excel_file_path())
    #excel.set_excel_file_path("e:\\sample.xlsx")
    #print(excel.get_excel_file_path())
    #print(excel.get_sheet_names())
    #excel.create_sheet_by_name("Sheet1")
    #excel.create_sheet_by_name("测试数据")
    #print(excel.get_sheet_names())
    #excel.save()
    #excel.set_sheet("不存在的sheet名称")
    #excel.set_sheet("测试数据")
    #print(excel.get_current_sheet_name())
    #print(excel.get_all_row_objects())
    #excel.set_sheet("不存在")
    excel = Excel("e:\\sample.xlsx")
    excel.set_sheet("测试数据")
    #print(excel.get_all_row_objects())
    #print(excel.get_one_row(100))
    #print(excel.get_one_row(3))
    #print(excel.get_all_cell_objects())
    #print(excel.get_all_cell_values())
    # print(excel.get_one_row_value(2))
    # print(excel.get_all_col_objects())
    # print(excel.get_all_col_values())
    # print(excel.get_one_col_object(2))
    # print(excel.get_one_col_values(2))
    # print(excel.get_cell_value(1, 1))
    # print(excel.get_cell_value(4, 3))
    # print(excel.get_cell_value(19, 4))
    # excel.write_lines(([11,22,33,44,55],))
    # excel.write_one_line(["我们1", "努力2", "每一天3"])
    # excel.write_cell(10, 10, "天气真好")
    # excel.write_cell(9, 10, "天气真好99",font="微软雅黑",color = "FF0000",pattern="00FF00")
    # excel.write_one_line(["我们1", "努力2", "每一天3"], "微软雅黑", "FF0000", "00FF00")
    # excel.write_one_line(["我们1", "努力2", "每一天3"])
    # excel.write_cell(17, 17, "失败")
    # excel.write_cell(18, 18, "fail")
    # excel.write_cell(19, 19, "成功")
    excel.write_lines([["表头1", "表头2", "表头3"], ["我们", "努力", "每一天"]], head_color="00FF00")
    excel.write_lines([["表头4", "表头5", "表头6"], ["我们", "努力", "每一天"]])

    excel.save()
