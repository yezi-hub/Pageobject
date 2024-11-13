import openpyxl
import time
import os
from openpyxl import load_workbook

class Excel:

    def __init__(self,excel_file_path):
        if not os.path.exists(excel_file_path):
            self.excel_file_path = None
            self.sheet = None
            self.wb = None
        else:
            self.excel_file_path = excel_file_path
            self.wb = load_workbook(self.excel_file_path)
            self.sheet = None

    def get_excel_file_path(self):
        return self.excel_file_path

    def set_excel_file_path(self,excel_file_path):
        if not os.path.exists(excel_file_path):
            self.excel_file_path = None
            self.wb = None
            self.sheet = None
        else:
            self.excel_file_path = excel_file_path
            self.sheet = None
            self.wb = load_workbook(self.excel_file_path)

    def get_sheet_names(self):
        if self.wb:
            return self.wb.sheetnames
        else:
            return []

if __name__ =="__main__":
    excel = Excel("e:\\aaaa.xlsx")
    print(excel.get_excel_file_path())
    excel = Excel("e:\\a.xlsx")
    print(excel.get_excel_file_path())
    excel.set_excel_file_path("e:\\sample.xlsx")
    print(excel.get_excel_file_path())
    print(excel.get_sheet_names())